from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(r'D:\Other\做账\5、京东2月数据-商家对账单.xlsx')

ws = wb.active

r = ws['A1:E12']

print(r)

